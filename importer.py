import openpyxl
import calendar
import re
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from models import MonthData

def import_from_excel(main_window):
    """
    Import schedule data from an Excel file into the current application state.
    Attempts to match People and Services by name/short_name.
    """
    path, _ = QFileDialog.getOpenFileName(
        main_window, 
        "Import from Excel", 
        "", 
        "Excel Files (*.xlsx *.xls)"
    )
    
    if not path:
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True) # data_only=True gets values, not formulas
        ws = wb.active
    except Exception as e:
        QMessageBox.critical(main_window, "Error", f"Could not open Excel file:\n{str(e)}")
        return

    # 1. Attempt to parse Month/Year from Cell A1 (Expected format: "JANUARY 2026")
    header_val = ws.cell(row=1, column=1).value
    if not header_val or not isinstance(header_val, str):
        choice = QMessageBox.question(
            main_window,
            "Unknown Date",
            "Could not detect Month/Year in cell A1.\n"
            "Assume it matches the currently selected month in the app?",
            QMessageBox.Yes | QMessageBox.No
        )
        if choice == QMessageBox.No:
            return
        
        target_year = int(main_window.year_combo.currentText())
        target_month = main_window.month_combo.currentIndex() + 1
    else:
        # Try to parse "MONTH YEAR"
        parts = header_val.strip().split()
        if len(parts) >= 2:
            month_str = parts[0]
            year_str = parts[-1]
            
            # Resolve Month
            target_month = None
            for i, name in enumerate(calendar.month_name):
                if name.upper() == month_str.upper():
                    target_month = i
                    break
            
            if not target_month:
                 # Try french names just in case since user seems French
                french_months = ["", "JANVIER", "FÉVRIER", "MARS", "AVRIL", "MAI", "JUIN", 
                                "JUILLET", "AOÛT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DÉCEMBRE"]
                if month_str.upper() in french_months:
                    target_month = french_months.index(month_str.upper())

            # Resolve Year
            try:
                target_year = int(year_str)
            except ValueError:
                target_year = None
        else:
            target_month = None
            target_year = None

        # Fallback if parsing failed
        if not target_month or not target_year:
            target_year = int(main_window.year_combo.currentText())
            target_month = main_window.month_combo.currentIndex() + 1
    
    # Check bounds
    if target_year < 2000 or target_year > 2100 or target_month < 1 or target_month > 12:
        QMessageBox.warning(main_window, "Invalid Date", f"Parsed date seems invalid: {target_month}/{target_year}")
        return

    # 2. Confirm matching context
    # We want to merge INTO the main_window state.
    # Identifying the MonthData object
    if (target_year, target_month) not in main_window.schedule:
        # Create it if it doesn't exist (e.g. importing a future month)
        main_window.schedule[(target_year, target_month)] = MonthData(target_year, target_month)
    
    month_data = main_window.schedule[(target_year, target_month)]
    
    # 3. Helpers for Fuzzy Matching
    
    def find_person(name_in_excel):
        """
        Attempts to match a name from Excel to a Person object.
        Excel might have: "Dupont Jean 100%" or "Jean DUPONT"
        """
        cleaned_name = str(name_in_excel).strip()
        
        # Remove percentage if present (e.g. "Name 80%")
        # Regex: look for digits followed by % at the end
        cleaned_name = re.sub(r'\s+\d+%$', '', cleaned_name)
        cleaned_name = cleaned_name.strip().lower()
        
        for person in main_window.people:
            # Check 1: Display Name match
            p_display = person.display_name.lower()
            if cleaned_name == p_display:
                return person
            
            # Check 2: formatted "Nom Prenom" or "Prenom Nom" containment
            p_nom = person.nom.lower()
            p_prenom = person.prenom.lower()
            
            # If Excel has "Dupont", match person(nom="Dupont")
            if cleaned_name == p_nom:
                return person
            
            # If Excel has "Jean", match person(prenom="Jean") - RISK of collision if multiple Jeans
            # Skipping overly simple single-name matches for safety unless unique?
            # Let's stick to composite matches
            
            if p_nom in cleaned_name and p_prenom in cleaned_name:
                return person
                
        return None

    def find_service(cell_value):
        """Matches short codes or full names."""
        if not cell_value:
            return None
        val = str(cell_value).strip()
        
        for s in main_window.services:
            if s.short_name.lower() == val.lower():
                return s.id
            if s.name.lower() == val.lower():
                return s.id
        return None

    # 4. Iterate Rows
    # In exporter.py, data starts at row 3
    # Rows loop until an empty first cell or reasonable limit
    
    changes_count = 0
    people_matched = 0
    
    # Determine columns range
    # In exporter: col 1=Name, col 2=Ratio. Days start at col 3 (but shifted by n_prev_days?)
    # Exporter Logic:
    # start_col = self.n_prev_days
    # But in Excel, Column C (3) is usually Day 1 if n_prev_days was 0.
    # Wait, the exporter includes previous days columns if n_prev_days > 0.
    # This makes mapping day-columns tricky if we don't know n_prev_days used at export time.
    # STRATEGY: Look at Row 2 (header). It contains the day numbers. Match column to day number!
    
    candidate_header_row = None
    day_map = {} # column_index -> day_number

    # Search rows 1 to 10 for day numbers
    best_row_count = 0
    
    for r in range(1, 11):
        possible_map = {}
        for col in range(1, 50): # Wider search (start at 1)
             val = ws.cell(row=r, column=col).value
             if val is None:
                 continue

             # Attempt to parse integer safely (handles "1", 1, 1.0, " 1 ")
             try:
                 # Convert to float first to handle "1.0" -> 1
                 val_float = float(val)
                 val_int = int(val_float)
                 
                 # Check if it's actually an integer (1.0 vs 1.5)
                 if abs(val_float - val_int) < 0.001:
                     if 1 <= val_int <= 31:
                         possible_map[col] = val_int
             except (ValueError, TypeError):
                 pass
        
        # We need a significant number of days to consider this a header row
        # (e.g. at least 5 days found)
        count = len(possible_map)
        if count > 5 and count > best_row_count:
             best_row_count = count
             candidate_header_row = r
             day_map = possible_map

    if not day_map or not candidate_header_row:
        # Debug info for user
        msg = "Could not find a row with day numbers (1..31).\nChecked rows 1-10, cols 1-50."
        QMessageBox.warning(main_window, "Format Error", msg)
        return

    # Iterate data rows starting after the header
    start_row = candidate_header_row + 1
    # Check if user has "Notes" or extra headers. Just iterate all the way down.
    # We will filter by valid Person name in column 1.

    for row in range(start_row, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=1).value
        if not name_cell:
            continue
            
        person = find_person(name_cell)
        if not person:
            # Skipping unknown person (e.g. Sections)
            # print(f"Skipping unknown person: {name_cell}")
            continue
            
        people_matched += 1
        
        # Iterate day columns
        for col, day_num in day_map.items():
            cell_val = ws.cell(row=row, column=col).value
            
            service_id = find_service(cell_val)
            
            # Apply to data
            # Logic: 
            # - If cell is empty in Excel -> Do we clear the assignment? 
            #   User said "adapt the import... work back and forth". 
            #   If Excel defines the schedule, empty probably means empty.
            #   BUT "holidays" or "weekends" might be marked visually but have no text.
            #   Ideally, we only update if we find a service token. 
            #   However, if they deleted a shift in Excel, we want it deleted here.
            #   Let's assume text presence = truth.
            
            # If valid service found -> Set it.
            if service_id:
                # Use canonical mutation to trigger rules, etc.
                # However, for BULK updates like this, calling the full chain including GUI refresh 
                # for every cell is too slow.
                # So we update MonthData directly, but then trigger A SINGLE global refresh.
                month_data.set_service(person.id, day_num, service_id)
                changes_count += 1
            
            # If cell has text but NO matching service -> Set "Unknown" service
            elif cell_val is not None and str(cell_val).strip() != "":
                 # Look for our special hidden service
                 unknown_service = next((s for s in main_window.services if s.id == "unknown"), None)
                 if unknown_service:
                     month_data.set_service(person.id, day_num, "unknown")
                     changes_count += 1
            
    # 5. Finish
    main_window.app_state.save_app_state(main_window.controller.to_dict())
    main_window.recompute_current_month_violations()
    main_window.refresh_row_headers()
    main_window.table.viewport().update()
    
    QMessageBox.information(
        main_window, 
        "Import Complete", 
        f"Processed {people_matched} people.\n"
        f"Updated {changes_count} assignments for {calendar.month_name[target_month]} {target_year}."
    )
    
    # Check if we are ALREADY viewing the target month
    current_y = int(main_window.year_combo.currentText())
    current_m = main_window.month_combo.currentIndex() + 1
    
    if (target_year == current_y) and (target_month == current_m):
        # We are on the same page, but we just bulk-updated the data.
        # We must rebuild the table to show the new assignments.
        main_window.recompute_current_month_violations()
        main_window.table_rebuilder.rebuild_cells() # Lighter than full finalize
        main_window.refresh_row_headers()
        main_window.table.viewport().update()

    # If we imported into a different month, offer to switch
    
    if (target_year != current_y) or (target_month != current_m):
        switch = QMessageBox.question(
            main_window,
            "Switch View?",
            f"Imported data was for {calendar.month_name[target_month]} {target_year}.\n"
            "Switch view to that month?",
            QMessageBox.Yes | QMessageBox.No
        )
        if switch == QMessageBox.Yes:
            main_window.year_combo.setCurrentText(str(target_year))
            main_window.month_combo.setCurrentIndex(target_month - 1)
            main_window.table_rebuilder.finalize() # Full refresh needed
            main_window.recompute_current_month_violations()
            main_window.refresh_row_headers()

