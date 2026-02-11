import openpyxl
import calendar
from openpyxl.styles import PatternFill, Alignment
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtGui import QImage, QPainter, QColor, QFont, QPen, QBrush
from PyQt5.QtCore import Qt, QRect, QRectF

from cell_authority import resolve_cell_appearance


def export_to_excel(self):
        month = self.month_combo.currentIndex() + 1
        year = int(self.year_combo.currentText())

        key = (year, month)
        if key not in self.schedule:
            print("No data to export for this month")
            return

        month_data = self.schedule[key]

        path, _ = QFileDialog.getSaveFileName(self, "Export to Excel", "", "Excel Files (*.xlsx)")
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{calendar.month_name[month].upper()} {year}"

        weekend_fill = PatternFill(
            start_color="DDDDDD",
            end_color="DDDDDD",
            fill_type="solid"
        )

        holiday_fill = PatternFill(
            start_color="DDDDDD",
            end_color="DDDDDD",
            fill_type="solid"
        )


        # First cell = MONTH year
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=2)
        ws.cell(row=1, column=1, value=f"{calendar.month_name[month].upper()} {year}")
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
        
        days_in_month = calendar.monthrange(year, month)[1]
        total_days = days_in_month + self.n_prev_days
        start_col = self.n_prev_days
        last_col = total_days - 1

        # Row 1 days' short names
        for col in range(start_col, total_days):
            m_data, day = self._resolve_day_context(col)
            weekday_index = calendar.weekday(m_data.year, m_data.month, day)
            weekday_short = self.table.FRENCH_DAYS[weekday_index]

            cell = ws.cell(row=1, column=col - start_col + 3)
            cell.value = weekday_short
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 2 days numbers
        for col in range(start_col, total_days):
            _, day = self._resolve_day_context(col)
            cell = ws.cell(row=2, column=col - start_col + 3)
            cell.value = day
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Row 1-2 Header for Notes
        notes_header_col = total_days - start_col + 3
        notes_header_cell = ws.cell(row=1, column=notes_header_col, value="NOTES")
        notes_header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=1, start_column=notes_header_col, end_row=2, end_column=notes_header_col)

        max_ratio_length = 0
        ws.freeze_panes = "C3"

        excel_row = 3

        # Fill rows with services
        for row_desc in self.rows:
            # Section row
            if row_desc["type"] == "section":
                # Write label in first column
                cell = ws.cell(row=excel_row, column=1, value=row_desc["label"])

                # Merge two first cells
                ws.merge_cells(
                    start_row=excel_row,
                    start_column=1,
                    end_row=excel_row,
                    end_column=2
                )
                # Merge across all day columns
                ws.merge_cells(
                    start_row=excel_row,
                    start_column=3,
                    end_row=excel_row,
                    end_column=last_col
                )

                # Center the label
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
                excel_row += 1
                continue
            
            # Person row
            if row_desc["type"] == "person":
                person_id = row_desc["person_id"]
                person = next(p for p in self.people if p.id == person_id)

                # Column A: display_name with percentage (only if not 100%)
                if person.percentage != 100:
                    name_display = f"{person.display_name}  {person.percentage}%"
                else:
                    name_display = person.display_name
                ws.cell(row=excel_row, column=1, value=name_display)

                # Column B : worked hours ratio
                summary = self.workload.monthly_summary(person, year, month)
                worked_hours = summary.worked
                total_hours = summary.expected

                ratio_float = 0 if total_hours == 0 else worked_hours / total_hours
                ratio_str = f"{worked_hours:g}h / {total_hours:g}h"

                cell_b = ws.cell(row=excel_row, column=2, value=ratio_str)
                cell_b.alignment = Alignment(horizontal="center", vertical="center")

                # ----- CONDITIONAL FORMATTING -----
                if ratio_float < 0.9:
                    fill_color = "ADD8FF"  # Light Blue
                elif ratio_float > 1.1:
                    fill_color = "FFB4B4"  # Light Red
                else:
                    fill_color = "B4E6B4"  # Light Green


                ws.cell(row=excel_row, column=1).fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )
                cell_b.fill = PatternFill(
                    start_color=fill_color,
                    end_color=fill_color,
                    fill_type="solid"
                )

                max_ratio_length = max(max_ratio_length, len(ratio_str))

                for col_offset, col in enumerate(range(start_col, total_days)):
                    month_data, day = self._resolve_day_context(col)
                    cell = ws.cell(row=excel_row, column=col_offset + 3)

                    service_id = month_data.get_service(person.id, day)

                    is_weekend = calendar.weekday(
                        month_data.year,
                        month_data.month,
                        day
                    ) >= 5

                    is_holiday = day in month_data.holidays

                    # --------------------------------------------------
                    # Cell authority decision (single source of truth)
                    # --------------------------------------------------
                    appearance = resolve_cell_appearance(
                        service_id, 
                        is_holiday, 
                        is_weekend, 
                        self.services
                    )

                    if appearance.type == "service" and appearance.service:
                        cell.value = appearance.service.short_name
                        cell.fill = PatternFill(
                            start_color=appearance.service.color_hex.strip("#"),
                            end_color=appearance.service.color_hex.strip("#"),
                            fill_type="solid"
                        )
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif appearance.type == "holiday":
                        cell.value = ""
                        cell.fill = holiday_fill
                    elif appearance.type == "weekend":
                        cell.value = ""
                        cell.fill = weekend_fill
                    else:  # empty
                        cell.value = ""
                    
                # Notes content
                comment = month_data.get_comment(person.id)
                ws.cell(row=excel_row, column=notes_header_col, value=comment)

                excel_row += 1

        # Compute max length of names
        max_name_length = max(len(person.short_name) for person in self.people) + 5

        # set column widths
        ws.column_dimensions['A'].width = max_name_length
        ws.column_dimensions['B'].width = max_ratio_length

        wb.save(path)
        print(f"Exported schedule to {path}")

def export_to_image(self):
    month = self.month_combo.currentIndex() + 1
    year = int(self.year_combo.currentText())

    key = (year, month)
    if key not in self.schedule:
        print("No data to export for this month")
        return

    path, _ = QFileDialog.getSaveFileName(self, "Export to Image", "", "PNG Image (*.png);;JPEG Image (*.jpg);;All Files (*)")
    if not path:
        return

    # A4 Landscape at 300 DPI
    # Width: 297mm * (300 / 25.4) ≈ 3508 px
    # Height: 210mm * (300 / 25.4) ≈ 2480 px
    IMG_WIDTH = 3508
    IMG_HEIGHT = 2480
    MARGIN = 50

    image = QImage(IMG_WIDTH, IMG_HEIGHT, QImage.Format_ARGB32)
    image.fill(Qt.white)

    painter = QPainter(image)
    painter.setRenderHint(QPainter.Antialiasing)

    # Fonts
    font_title = QFont("Arial", 40, QFont.Bold)
    font_header = QFont("Arial", 14, QFont.Bold)
    font_cell = QFont("Arial", 12)
    font_bold_cell = QFont("Arial", 12, QFont.Bold)

    # --- Metrics ---
    days_in_month = calendar.monthrange(year, month)[1]
    
    # We want: Name Col | Day Cols ... | Stats Col
    # Let's assign proportions or fixed sizes.
    # Name need decent space. Stats need small space. Days need the rest.
    
    # Available width
    draw_width = IMG_WIDTH - 2 * MARGIN
    draw_height = IMG_HEIGHT - 2 * MARGIN
    
    # 15% for Name, 10% for Stats, 75% for Days
    col_name_width = draw_width * 0.10
    col_stats_width = draw_width * 0.10
    col_day_width = (draw_width - col_name_width - col_stats_width) / days_in_month

    # --- Draw Title ---
    french_months = [
        "", "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ]
    month_name_fr = french_months[month]

    painter.setFont(font_title)
    title_rect = QRect(MARGIN, MARGIN, int(draw_width), 80)
    painter.drawText(title_rect, Qt.AlignCenter, f"{month_name_fr.upper()} {year}")

    current_y = MARGIN + 100

    # --- Draw Headers ---
    # We have 2 header rows: Day Names (Lun, Mar...) and Day Numbers (1, 2...)
    # But first, calculate row height based on content to fit?
    
    # Count rows: Header (2) + Data Rows (Sections + People)
    num_data_rows = len(self.rows)
    total_rows = num_data_rows + 2 
    
    # Remaining height for table
    table_height = IMG_HEIGHT - MARGIN - current_y
    
    # Calculate row height
    # use a comfortable height, but scale down if too many people
    target_row_height = 45
    min_row_height = 30
    
    # Check if it fits
    if target_row_height * total_rows > table_height:
        actual_row_height = max(min_row_height, int(table_height / total_rows))
    else:
        actual_row_height = target_row_height

    header_height = actual_row_height * 2

    painter.setFont(font_header)
    painter.setPen(QPen(Qt.black, 2))

    # Name Header
    rect_name = QRectF(MARGIN, current_y, col_name_width, header_height)
    painter.drawRect(rect_name)
    painter.drawText(rect_name, Qt.AlignCenter, f"{month_name_fr.upper()} {year}")

    # Day Headers
    french_days = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    
    for day in range(1, days_in_month + 1):
        x = MARGIN + col_name_width + (day - 1) * col_day_width
        
        # Determine day name
        weekday = calendar.weekday(year, month, day)
        day_str = french_days[weekday]
        
        # Check weekend/holiday for background
        is_weekend = weekday >= 5
        is_holiday = day in self.schedule[key].holidays
        
        bg_brush = Qt.NoBrush
        if is_weekend or is_holiday:
            bg_brush = QBrush(QColor("#DDDDDD"))
            painter.fillRect(QRectF(x, current_y, col_day_width, header_height), bg_brush)
        
        # Top half: Name
        r_top = QRectF(x, current_y, col_day_width, actual_row_height)
        painter.drawRect(r_top)
        painter.drawText(r_top, Qt.AlignCenter, day_str)
        
        # Bottom half: Number
        r_bot = QRectF(x, current_y + actual_row_height, col_day_width, actual_row_height)
        painter.drawRect(r_bot)
        painter.drawText(r_bot, Qt.AlignCenter, str(day))

    # Stats Header
    x_stats = MARGIN + col_name_width + days_in_month * col_day_width
    rect_stats = QRectF(x_stats, current_y, col_stats_width, header_height)
    painter.drawRect(rect_stats)
    painter.drawText(rect_stats, Qt.AlignCenter, "BILAN")

    current_y += header_height

    # --- Draw Rows ---
    painter.setFont(font_cell)
    
    for row_data in self.rows:
        x = MARGIN
        
        if row_data["type"] == "section":
            # Section Row
            painter.setBrush(QBrush(QColor("#E0E0E0")))
            painter.drawRect(QRectF(x, current_y, draw_width, actual_row_height))
            painter.setBrush(Qt.NoBrush)
            
            painter.setFont(font_header)
            painter.drawText(QRectF(x, current_y, draw_width, actual_row_height), Qt.AlignCenter, row_data["label"])
            painter.setFont(font_cell)
            
        elif row_data["type"] == "person":
            person = next(p for p in self.people if p.id == row_data["person_id"])
            summary = self.workload.monthly_summary(person, year, month)
            
            # --- Name Cell ---
            # Background based on ratio
            ratio = summary.ratio
            if ratio < 0.9:
                name_bg = QColor("#ADD8FF")
            elif ratio > 1.1:
                name_bg = QColor("#FFB4B4")
            else:
                name_bg = QColor("#B4E6B4")
            
            painter.setBrush(QBrush(name_bg))
            painter.drawRect(QRectF(x, current_y, col_name_width, actual_row_height))
            painter.setBrush(Qt.NoBrush)
            
            # Text
            name_text = person.display_name
            if person.percentage != 100:
                name_text += f" ({person.percentage}%)"
            
            # Draw name with some padding
            painter.drawText(QRectF(x + 5, current_y, col_name_width - 10, actual_row_height), Qt.AlignVCenter | Qt.AlignLeft, name_text)
            
            # --- Day Cells ---
            cur_x = x + col_name_width
            month_data = self.schedule[key]
            
            for day in range(1, days_in_month + 1):
                weekday = calendar.weekday(year, month, day)
                is_weekend = weekday >= 5
                is_holiday = day in month_data.holidays
                
                service_id = month_data.get_service(person.id, day)
                
                # Resolve appearance
                appearance = resolve_cell_appearance(service_id, is_holiday, is_weekend, self.services)
                
                cell_bg = Qt.white
                cell_text = ""
                
                if appearance.type == "service" and appearance.service:
                    cell_bg = QColor(appearance.service.color_hex)
                    cell_text = appearance.service.short_name
                elif appearance.type == "holiday":
                    cell_bg = QColor("#DDDDDD")
                elif appearance.type == "weekend":
                    cell_bg = QColor("#DDDDDD")
                    
                painter.setBrush(QBrush(cell_bg))
                painter.drawRect(QRectF(cur_x, current_y, col_day_width, actual_row_height))
                painter.setBrush(Qt.NoBrush)
                
                if cell_text:
                    painter.drawText(QRectF(cur_x, current_y, col_day_width, actual_row_height), Qt.AlignCenter, cell_text)
                
                cur_x += col_day_width
                
            # --- Stats Cell ---
            painter.setBrush(QBrush(name_bg)) # Match name cell bg
            painter.drawRect(QRectF(cur_x, current_y, col_stats_width, actual_row_height))
            painter.setBrush(Qt.NoBrush)
            
            stats_text = f"{summary.worked:g} / {summary.expected:g}"
            painter.drawText(QRectF(cur_x, current_y, col_stats_width, actual_row_height), Qt.AlignCenter, stats_text)
            
        current_y += actual_row_height

    painter.end()
    
    if image.save(path):
        print(f"Exported image to {path}")
    else:
        print(f"Failed to save image to {path}")
